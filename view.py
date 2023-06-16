'''
This module contains the View class which contains all the visual aspects of the system.
'''

# Imports
import tkinter as tk
import json
from widgets_for_gui import *
from observer_subject import Observer
from model_hypgeo import *
from openpyxl import load_workbook, Workbook

# Classes
class View(tk.Tk, Observer):
    '''
    Class that manages all visual aspects of the program.
    '''
    def __init__(self, model, controller):
        
        # Interaction with the model and the controller.
        self.model = model              # Which model should be observed.
        self.model.attach(self)         # Attach to this model.
        self.controller = controller    # Controller contains all event handlers.

        # Count/Storing variables.
        self.groups = {}
        self.key = 0

        # Setup a window.
        tk.Tk.__init__(self)
        self.geometry("540x520")
        self.resizable(False, True)
        self.title("Drawing Probability Master v2")      
        self.init_ui()                                  # Widgets.
        self.eval('tk::PlaceWindow . center')           # Center the window.
        self.bind("<1>", self.set_focus)                # Allways focus on the widget on left click.
        self.mainloop()

    def init_ui(self):
        '''
        Creates all widgets.
        '''
        # Widget sizes.
        self.label_width = 10
        self.text_size = 10
        self.entry_width = 5

        # Main frames: Create, configure and arragne with grid.
        self.main_frm_top = tk.Frame(master=self, relief=tk.RIDGE, borderwidth=5, width=520, height=50)
        self.main_frm_top.grid_propagate(0)
        self.main_frm_bottom = Scrollable_Frame(master=self, relief=tk.RIDGE, borderwidth=5)

        tk.Grid.rowconfigure(self=self, index=1, weight=1)

        self.main_frm_top.grid(row=0, column=0, padx=10, pady=5, sticky="nesw")
        self.main_frm_bottom.grid(row=1, column=0, padx=10, pady=5, sticky="nesw")
        
        # Top labels, buttons, entries: Create, configure/bind and arrange with grid.
        self.lbl_deck_size = tk.Label(master=self.main_frm_top, text="Deck Size:", font=("Helvetica", self.text_size, "bold"), anchor="e")
        self.lbl_sample_size = tk.Label(master=self.main_frm_top, text="Sample Size:", font=("Helvetica", self.text_size, "bold"), anchor="e")

        self.btn_calculate = tk.Button(master=self.main_frm_top, text="CALCULATE", font=("Helvetica", self.text_size), anchor="center", command=self.controller.on_calculate)

        self.ent_deck_size = tk.Entry(master=self.main_frm_top, font=("Helvetica", self.text_size), width=self.entry_width, validate="key")
        self.ent_deck_size.configure(validatecommand=(self.ent_deck_size.register(self.validate),'%d', '%P'))
        self.ent_deck_size.bind('<FocusOut>', lambda event: self.controller.on_deck_size(self.ent_deck_size.get()))
        self.ent_sample_size = tk.Entry(master=self.main_frm_top, font=("Helvetica", self.text_size), width=self.entry_width, validate="key")
        self.ent_sample_size.configure(validatecommand=(self.ent_sample_size.register(self.validate),'%d', '%P'))
        self.ent_sample_size.bind('<FocusOut>', lambda event: self.controller.on_sample_size(self.ent_sample_size.get()))

        self.lbl_deck_size.grid(row=0, column=0, padx=5, pady=5, sticky="nesw")
        self.lbl_sample_size.grid(row=0, column=2, padx=5, pady=5, sticky="nesw")

        self.btn_calculate.grid(row=0, column=4, padx=40, pady=5, sticky="nesw")

        self.ent_deck_size.grid(row=0, column=1, padx=5, pady=5, sticky="nesw")
        self.ent_sample_size.grid(row=0, column=3, padx=5, pady=5, sticky="nesw")

        # Bottom widgets: Create, configure and arrange with grid
        self.lbl_group_name = tk.Label(master=self.main_frm_bottom.get_frm_container(), text="Group Name\n(optional)", font=("Helvetica", self.text_size, "bold"), anchor="center", width=self.label_width)
        self.lbl_group_success_deck = tk.Label(master=self.main_frm_bottom.get_frm_container(), text="Success in\nDeck", font=("Helvetica", self.text_size, "bold"), anchor="center", width=self.label_width)
        self.lbl_group_min_success_sample = tk.Label(master=self.main_frm_bottom.get_frm_container(), text="Min. Success\nin Sample", font=("Helvetica", self.text_size, "bold"), anchor="center", width=self.label_width)
        self.lbl_group_max_success_sample = tk.Label(master=self.main_frm_bottom.get_frm_container(), text="Max. Success\nin Sample", font=("Helvetica", self.text_size, "bold"), anchor="center", width=self.label_width)
        self.lbl_group_delete = tk.Label(master=self.main_frm_bottom.get_frm_container(), text="Delete\nGroup", font=("Helvetica", self.text_size, "bold"), anchor="center", width=self.label_width)

        self.btn_add_group = tk.Button(master=self.main_frm_bottom.get_frm_container(), text="ADD GROUP", font=("Helvetica", self.text_size), anchor="center", command=lambda: [self.add_group_view(), self.controller.on_add_group()])
        
        self.lbl_group_name.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.lbl_group_success_deck.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.lbl_group_min_success_sample.grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.lbl_group_max_success_sample.grid(row=0, column=3, padx=5, pady=5, sticky="w")
        self.lbl_group_delete.grid(row=0, column=4, padx=5, pady=5, sticky="w")

        self.btn_add_group.grid(row=1, column=0, columnspan=5, padx=5, pady=5, sticky="nesw")

    def add_group_view(self):
        '''
        Add a new group to the display.
        '''
        key = self.key
        self.groups[key] = []
        group = self.groups[key]

        group.append(tk.Frame(master=self.main_frm_bottom.get_frm_container()))   # Frame for one grouo
        group.append(tk.Entry(master=group[0], font=("Helvetica", self.text_size), width=(2 * self.entry_width)))
        group.append(tk.Entry(master=group[0], font=("Helvetica", self.text_size), width=(self.entry_width), validate="key"))
        group.append(tk.Entry(master=group[0], font=("Helvetica", self.text_size), width=(self.entry_width), validate="key"))
        group.append(tk.Entry(master=group[0], font=("Helvetica", self.text_size), width=(self.entry_width), validate="key"))
        group.append(tk.Button(master=group[0], text=" x ", font=("Helvetica", self.text_size, "bold"), anchor="center", command=lambda key=key: [self.del_group_view(key), self.controller.on_del_group(key)]))

        for index, command in zip([2,3,4], [self.controller.on_group_size, self.controller.on_group_min, self.controller.on_group_max]):
            entry = group[index]
            entry.configure(validatecommand=(entry.register(self.validate),'%d', '%P'))                                 # Only allow integers as inputs for all entries.
            entry.bind('<FocusOut>', lambda event, key=key, command=command, entry=entry: command(key, entry.get()))    # Bind controller method to every entry.

        group[0].grid(row=(key + 1), column=0, columnspan=5)             # Arrange the group frame.
        group[1].grid(row=0, column=0, padx=5, pady=5, sticky="ns")
        for column in range(2, 6, 1):
            group[column].grid(row=0, column=column, padx=30, pady=5, sticky="ns")  # Arrange the group widgets.

        self.btn_add_group.grid_forget()                                            # Arrange the button at the bottom of all groups
        self.btn_add_group.grid(row=(key + 2), column=0, columnspan=5, padx=5, pady=5, sticky="nesw")

        self.key += 1

    def del_group_view(self, key):
        '''
        Remove a group from the display.
        '''
        for widget in self.groups[key]:
            widget.destroy()
        self.groups.pop(key)

    def validate(self, type_of_action, entry_value):
        '''
        Validate function that only allows integers as inserts in entries.
        '''
        if type_of_action == '1':           # '1' is insert
            if not entry_value.isdigit():
                return False
            else:
                return True
        else:
            return True
        
    def set_focus(self, event=None):
        '''
        Focuses on the current widget.
        '''
        x, y = self.winfo_pointerxy()
        self.winfo_containing(x, y).focus()

    def update(self, update_event, **kwargs):
        '''
        Update the GUI.
        '''
        if update_event == "start calculate":
            pass

        elif update_event == "end calculte":
            self.popup_result()
        
        elif update_event == "invalid deck size":
            invalid_deck_size = int(self.ent_deck_size.get())
            sample_size = self.model.get_sample_size()

            if invalid_deck_size <= sample_size:
                self.ent_sample_size.delete(0, "end")
                self.ent_sample_size.insert(0, invalid_deck_size)
                self.controller.on_sample_size(invalid_deck_size)
                self.controller.on_deck_size(invalid_deck_size)

            else:
                self.ent_deck_size.delete(0, "end")
                self.ent_deck_size.insert(0, self.model.get_deck_size())

        elif update_event == "invalid sample size":
            invalid_sample_size = int(self.ent_sample_size.get())
            deck_size = self.model.get_deck_size()

            if invalid_sample_size >= deck_size:
                self.ent_sample_size.delete(0, "end")
                self.ent_sample_size.insert(0, deck_size)
                self.controller.on_sample_size(deck_size)

            else:
                self.ent_sample_size.delete(0, "end")
                self.ent_sample_size.insert(0, self.model.get_sample_size())

        elif update_event == "invalid group":
            pass

        elif update_event == "invalid group size":
            key = kwargs["group_key"]
            invalid_size = int(self.groups[key][2].get())
            group_min = self.model.get_group_min(key)
            group_max = self.model.get_group_max(key)
            unassigned_cards = self.model.get_unassigned_cards()
            if invalid_size >= (unassigned_cards + self.model.get_group_size(key)):
                self.groups[key][2].delete(0, "end")
                self.groups[key][2].insert(0, unassigned_cards + self.model.get_group_size(key))
                self.controller.on_group_size(key, unassigned_cards + self.model.get_group_size(key))

            elif invalid_size <= group_max and invalid_size >= group_min:
                self.groups[key][4].delete(0, "end")
                self.groups[key][4].insert(0, invalid_size)
                self.controller.on_group_max(key, invalid_size)
                self.controller.on_group_size(key, invalid_size)

            elif invalid_size <= group_min:
                self.groups[key][3].delete(0, "end")
                self.groups[key][3].insert(0, invalid_size)
                self.groups[key][4].delete(0, "end")
                self.groups[key][4].insert(0, invalid_size)
                self.controller.on_group_min(key, invalid_size)
                self.controller.on_group_max(key, invalid_size)
                self.controller.on_group_size(key, invalid_size)

            else:
                self.groups[key][2].delete(0, "end")
                self.groups[key][2].insert(0, self.model.get_group_min(key))

        elif update_event == "invalid group min":
            key = kwargs["group_key"]
            invalid_min = int(self.groups[key][3].get())
            group_size = self.model.get_group_size(key)
            group_max = self.model.get_group_max(key)

            if invalid_min >= group_max and invalid_min <= group_size:
                self.groups[key][4].delete(0, "end")
                self.groups[key][4].insert(0, invalid_min)
                self.controller.on_group_max(key, invalid_min)
                self.controller.on_group_min(key, invalid_min)

            elif invalid_min >= group_max and invalid_min >= group_size:
                self.groups[key][4].delete(0, "end")
                self.groups[key][4].insert(0, group_size)
                self.groups[key][3].delete(0, "end")
                self.groups[key][3].insert(0, group_size)
                self.controller.on_group_max(key, group_size)
                self.controller.on_group_min(key, group_size)
            
            else:
                self.groups[key][3].delete(0, "end")
                self.groups[key][3].insert(0, self.model.get_group_min(key))

        elif update_event == "invalid group max":
            key = kwargs["group_key"]
            invalid_max = int(self.groups[key][4].get())
            group_size = self.model.get_group_size(key)
            group_min = self.model.get_group_min(key)

            if invalid_max <= group_min:
                self.groups[key][4].delete(0, "end")
                self.groups[key][4].insert(0, group_min)
                self.controller.on_group_max(key, group_min)

            elif invalid_max >= group_size:
                self.groups[key][4].delete(0, "end")
                self.groups[key][4].insert(0, group_size)
                self.controller.on_group_max(key, group_size)

            else:
                self.groups[key][4].delete(0, "end")
                self.groups[key][4].insert(0, self.model.get_group_max(key))   

        elif update_event == "key error":
            pass    

    def popup_result(self):
        '''
        Generates a popup in the middle of the application with "text".
        '''
        self.popup_result = tk.Toplevel()                    # Create a popup window.
        self.popup_result.geometry("390x70")                 # Set size.
        self.popup_result.resizable(False, False)            # Lock size.
        x = self.winfo_x()
        y = self.winfo_y()
        self.popup_result.geometry("+%d+%d" %(x+75,y+225))   # Center the popup in front of the main window.
        self.popup_result.grab_set()                         # "Freezes" the main window until the popup is closed.

        # Label
        text = ("The probability of this configuration is: " + str(format_float(self.model.get_result(), factor=100)) + "%")
        lbl_popup = tk.Label(master=self.popup_result, text=text, height=1, font=("Helvetica", "11"), anchor="center")
        lbl_popup.pack(padx=5, pady=5)

        # Button
        btn_export = tk.Button(master=self.popup_result, text="EXPORT TO EXCEL", font=("Helvetica", self.text_size), anchor="center", command=self.popup_export)
        btn_export.pack(padx=5, pady=5)

        self.popup_result.mainloop()

    def popup_export(self):
        '''
        Popup for exporting the current data to an excel file.
        '''
        with open("refrences.json", "r") as read_storage:
            path = json.load(read_storage)["excel path"]

        self.popup_result.destroy()
        self.popup_export = tk.Toplevel()                    # Create a popup window
        self.popup_export.geometry("390x120")                # Set size.
        self.popup_export.resizable(False, False)            # Lock size.
        x = self.winfo_x()
        y = self.winfo_y()
        self.popup_export.geometry("+%d+%d" %(x+75,y+200))   # Center the popup in front of the main window.
        self.popup_export.grab_set()                         # "Freezes" the main window until the popup is closed.

        # Labels
        lbl_path = tk.Label(master=self.popup_export, text="Path:", font=("Helvetica", self.text_size, "bold"), anchor="e")
        lbl_path.grid(row=0, column=0, padx=5, pady=5)

        # Entries
        ent_path = tk.Entry(master=self.popup_export, font=("Helvetica", self.text_size), width=9*self.entry_width)
        ent_path.insert("end", path)
        ent_path.grid(row=0, column=1, padx=5, pady=5)

        # Buttons
        frm_btn = tk.Frame(master=self.popup_export, borderwidth=0, width=390)
        frm_btn.grid(row=1, column=0, columnspan=2)
        btn_export_to_path = tk.Button(master=frm_btn, text="EXPORT", font=("Helvetica", self.text_size), anchor="center", width=12, command=lambda: self.on_export(ent_path.get(), self.get_data()))
        btn_create_at_path = tk.Button(master=frm_btn, text="CREATE NEW", font=("Helvetica", self.text_size), anchor="center", width=12, command=lambda: self.on_create_export(ent_path.get(), self.get_data()))
        btn_export_to_path.grid(row=0, column=0, padx=30)
        btn_create_at_path.grid(row=0, column=1, padx=30)

    def get_data(self):
        '''
        Get entire relevant data for the current calculation and format it, so it can be entered into an excel file.
        '''
        data = [[], 
                [], 
                ["deck size", self.model.get_deck_size()],
                ['sample size', self.model.get_sample_size()],
                []]
        
        model_defined_groups = self.model.get_defined_groups()
        for key in model_defined_groups:
            model_group = model_defined_groups[key]
            group_data = [[self.groups[key][1].get()],
                          ["cards in deck", model_group[0]],
                          ["min. in sample", model_group[1]],
                          ["max. in sample", model_group[2]],
                          []
                          ]
            data.extend(group_data)

        data.append(["probability", self.model.get_result()])

        return data

    def on_export(self, path, data):
        '''
        Load existing Excel file, add data.
        '''
        self.store_path(path)
        wb = load_workbook(path)
        self.export(path, data, wb)
        

    def on_create_export(self, path, data):
        '''
        Create new excel file, add data to it.
        '''
        self.store_path(path)
        wb = Workbook()
        self.export(path, data, wb)

    def export(self, path, data, wb):
        '''
        Adds data to excel.
        '''
        ws = wb[wb.sheetnames[0]]

        for row in data:
            ws.append(row)
        
        wb.save(path)
        wb.close()
        self.popup_export.destroy()

    def store_path(self, path):
        '''
        Stores last path in json.
        '''
        with open("refrences.json", "r") as read_storage:
            refrences = json.load(read_storage)

        refrences["excel path"] = path

        with open("refrences.json", "w") as write_storage:
            json.dump(refrences, write_storage, indent=4, separators=(",",":"))
